import openpyxl
import json

filename = 'upload/RPT_20260212_이지원재고수불표_SELF_완료.xlsx'
wb = openpyxl.load_workbook(filename, data_only=True)

sheets_data = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(r is not None for r in row):
            rows.append([str(r) if r is not None else '' for r in row])
    if len(rows) > 0:
        sheets_data[name] = rows

with open('upload/all_excel_sheets.json', 'w', encoding='utf-8') as f:
    json.dump(sheets_data, f, ensure_ascii=False, indent=2)

print(f"총 {len(sheets_data)}개 시트 파싱 완료 -> upload/all_excel_sheets.json")
